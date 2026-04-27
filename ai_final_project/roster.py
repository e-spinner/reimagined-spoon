"""Read student names from Excel (.xlsx) or LibreOffice Calc (.ods) roster files."""

from __future__ import annotations

from pathlib import Path


def read_student_names(path: Path) -> list[str]:
    """First column of the first sheet, non-empty rows, optional header row skipped."""
    # Route by extension so each parser can stay format-specific and simple.
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        names = _read_xlsx_column_a(path)
    elif suffix == ".ods":
        names = _read_ods_column_a(path)
    elif suffix == ".xls":
        raise ValueError(
            "Legacy .xls is not supported. Save the roster as .xlsx (Excel) or .ods (Calc)."
        )
    else:
        raise ValueError(f"Unsupported file type “{suffix}”. Use .xlsx, .xlsm, or .ods.")

    # Normalize imported names and remove likely header labels.
    names = _strip_header_if_present(names)
    # Drop exact duplicates while keeping order
    seen: set[str] = set()
    unique: list[str] = []
    for n in names:
        if n not in seen:
            seen.add(n)
            unique.append(n)
    return unique


def _strip_header_if_present(names: list[str]) -> list[str]:
    # Heuristic: many class rosters use a simple "Name"/"Student Name" header.
    # If detected in the first row, discard it before returning names.
    if not names:
        return names
    first = names[0].strip().lower()
    header_hints = frozenset(
        {
            "name",
            "student",
            "student name",
            "full name",
            "students",
            "last name",
        }
    )
    if first in header_hints or first.endswith(" name") or first.endswith(" names"):
        return names[1:]
    return names


def _read_xlsx_column_a(path: Path) -> list[str]:
    from openpyxl import load_workbook

    # `data_only=True` reads calculated values for formula cells.
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        # Contract: first worksheet, first column (A), non-empty text values.
        ws = wb.active
        out: list[str] = []
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            if not row:
                continue
            val = row[0]
            if val is None:
                continue
            s = str(val).strip()
            if s:
                out.append(s)
        return out
    finally:
        wb.close()


def _read_ods_column_a(path: Path) -> list[str]:
    from odf import teletype
    from odf.opendocument import load
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    doc = load(str(path))
    tables = doc.spreadsheet.getElementsByType(Table)
    if not tables:
        return []
    # Match Excel behavior: first sheet only.
    table = tables[0]
    out: list[str] = []
    for row in table.getElementsByType(TableRow):
        cells = row.getElementsByType(TableCell)
        if not cells:
            continue
        cell = cells[0]
        # ODS cells can contain multiple paragraph nodes; concatenate them.
        parts: list[str] = []
        for p in cell.getElementsByType(P):
            parts.append(str(teletype.extractText(p)))
        s = "".join(parts).strip()
        if s:
            out.append(s)
    return out
