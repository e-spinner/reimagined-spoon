from pathlib import Path
import sys
import unittest

import fitz
from PySide6.QtWidgets import QApplication


PROJECT_ROOT = Path(__file__).resolve().parents[1] / "reimagined-spoon"
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from ai_final_project.grading_extract import extract_submission_numeric_answer
from ai_final_project.ui.main_window import MainWindow


class TestMathOutputArtifacts(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])
        cls.submission_pdf = Path(__file__).resolve().parents[1] / "Unit Tests" / "fixtures" / "MichaelSmithHW1.pdf"
        cls.window = MainWindow()
        cls.output_dir = Path(__file__).resolve().parents[0] / "output" / "math_artifacts"
        cls.output_dir.mkdir(parents=True, exist_ok=True)

    def test_write_math_annotations_outputs_mark_and_score_summary(self) -> None:
        extracted = extract_submission_numeric_answer(self.submission_pdf)
        dst = self.output_dir / "graded_sample_correct.pdf"
        self.window._write_math_annotations(
            src_pdf=self.submission_pdf,
            dst_pdf=dst,
            extracted=extracted,
            awarded_points=5.0,
            max_points=5.0,
            percent=100.0,
            is_correct=True,
        )
        self.assertTrue(dst.exists())
        self.assertGreater(dst.stat().st_size, 0)
        with fitz.open(str(dst)) as doc:
            text = (doc[0].get_text() or "") if len(doc) else ""
        self.assertIn("5.00/5.00", text)
        self.assertIn("100.0%", text)
        (self.output_dir / "graded_sample_correct.summary.txt").write_text(
            "Expected summary:\n5.00/5.00 (100.0%)\n",
            encoding="utf-8",
        )

    def test_write_grades_spreadsheet_writes_second_column(self) -> None:
        rows = [("George Alan", "5.00 (100.0%)"), ("Michael Smith", "3.00 (60.0%)")]
        self.window._submissions_folder = self.output_dir
        self.window._roster_path = None
        out = self.window._write_grades_spreadsheet(rows)
        self.assertTrue(out.exists())
        self.assertEqual(out.name, "grades_output.xlsx")

        from openpyxl import load_workbook

        wb = load_workbook(out, data_only=True)
        try:
            ws = wb.active
            self.assertEqual(ws.cell(row=1, column=1).value, "Name")
            self.assertEqual(ws.cell(row=1, column=2).value, "Grade")
            self.assertEqual(ws.cell(row=2, column=2).value, "5.00 (100.0%)")
            self.assertEqual(ws.cell(row=3, column=2).value, "3.00 (60.0%)")
        finally:
            wb.close()


if __name__ == "__main__":
    unittest.main()
